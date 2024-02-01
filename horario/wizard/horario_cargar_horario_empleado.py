import base64
from datetime import *
from odoo.http import request
import openpyxl
import requests
import math
import pytz
import json
from io import BytesIO
from odoo import api, fields, models, _
from odoo.exceptions import ValidationError

class HorarioCargarHorarioEmpleadoWizard(models.TransientModel):
    _name = "horario.cargar.horario.empleado.wizard"
    _inherit = ["mail.thread", "mail.activity.mixin"]
    _description = "Cargar Horario de los empleados"

    documentoHorario= fields.Binary(
        string='Horario'
    )

    def cargar_horario_empleado(self):
        step = timedelta(1)
        ICPSudo = self.env['ir.config_parameter'].sudo()
        wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.documentoHorario)),read_only=True)

        ws = wb.active

        empleadosNoEncontradosList = []
        for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=None, values_only=True):
            semanas = 0
            empleados = self.env['hr.employee'].search(['&',('active','=',True),('department_id.name', '!=','Docentes'),('department_id','!=',"Inactivos")])

            empleado = list(filter(lambda x: (str(x.identification_id).replace("-", "") == str(record[3]).replace("-", "") ),empleados))

            if len(empleado) <= 0:
                dict = {
                    'identificacion':str(record[3]),
                    'motivo': 'No Se Encontro al Empleado',
                }
                empleadosNoEncontradosList.append(dict)
                continue

            horario = self.env['horario.empleado'].search([('empleado_id', '=', empleado[0].id)])

            if bool(horario) == False:
                valsEmpleado = {
                    'empleado_id': empleado[0].id
                }
                horario = self.env['horario.empleado'].sudo().create(valsEmpleado)

            if record[0] == 'NS':
                fechaFinal = record[4] + timedelta(weeks=int(record[1]))
                semanas = int(record[1]) - 1

            elif record[0] == 'Fijo':
                fechaFinal = date(record[4].year, 12, 31)
                semanas = (52 - record[4].isocalendar()[1])


            fechaHorario = record[4].date()
            fechaHorarioT = (record[4] + (timedelta(days=6))).date()
            semanasWhile = 0
            horarioSet = 0
            while semanasWhile <= semanas:

                if record[2] == 'SI':
                    if semanasWhile % 2 == 0:
                        horarioSet = 0
                    else:
                        horarioSet = 14

                diaHorario = list(filter(lambda x: (x.fechaDesde == fechaHorario),horario.horarioEmpleado_ids))
                if bool(diaHorario) == False:
                    vals = {
                        'name': 'Horario desde: '+ str(fechaHorario) + ' hasta: ' + str(fechaHorarioT),
                        'fechaDesde': fechaHorario,
                        'fechaHasta': fechaHorarioT,
                        'horarioEmpleado_id': horario.id,
                    }
                    diaHorario = self.env['horario.empleado.line'].sudo().create(vals)
                else:
                    diaHorario = diaHorario[0]

                while fechaHorario <= fechaHorarioT:

                    if fechaHorario.weekday() == 0:
                        if record[5 + horarioSet] != None:
                            diaHorario.horaInicioLunes = datetime.strptime(str(fechaHorario) + ' ' + str(record[5+horarioSet] + ':00'),"%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                            diaHorario.horaFinalLunes = datetime.strptime(str(fechaHorario) + ' ' + str(record[6+horarioSet]  + ':00'),"%Y-%m-%d %H:%M:%S") + timedelta(hours=6)

                    elif fechaHorario.weekday() == 1:
                        if record[7 + horarioSet] != None:
                            diaHorario.horaInicioMartes = datetime.strptime(str(fechaHorario) + ' ' + str(record[7+horarioSet] + ':00'),"%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                            diaHorario.horaFinalMartes = datetime.strptime(str(fechaHorario) + ' ' + str(record[8+horarioSet]  + ':00'),"%Y-%m-%d %H:%M:%S") + timedelta(hours=6)

                    elif fechaHorario.weekday() == 2:
                        if record[9 + horarioSet] != None:
                            diaHorario.horaInicioMiercoles = datetime.strptime(str(fechaHorario) + ' ' + str(record[9+horarioSet] + ':00'),"%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                            diaHorario.horaFinalMiercoles = datetime.strptime(str(fechaHorario) + ' ' + str(record[10+horarioSet]  + ':00'),"%Y-%m-%d %H:%M:%S") + timedelta(hours=6)

                    elif fechaHorario.weekday() == 3:
                        if record[11 + horarioSet] != None:
                            diaHorario.horaInicioJueves = datetime.strptime(str(fechaHorario) + ' ' + str(record[11+horarioSet] + ':00'),"%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                            diaHorario.horaFinalJueves = datetime.strptime(str(fechaHorario) + ' ' + str(record[12+horarioSet]  + ':00'),"%Y-%m-%d %H:%M:%S") + timedelta(hours=6)

                    elif fechaHorario.weekday() == 4:
                        if record[13 + horarioSet] != None:
                            diaHorario.horaInicioViernes = datetime.strptime(str(fechaHorario) + ' ' + str(record[13+horarioSet] + ':00'),"%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                            diaHorario.horaFinalViernes = datetime.strptime(str(fechaHorario) + ' ' + str(record[14+horarioSet]  + ':00'),"%Y-%m-%d %H:%M:%S") + timedelta(hours=6)

                    elif fechaHorario.weekday() == 5:
                        if record[15 + horarioSet] != None:
                            diaHorario.horaInicioSabado = datetime.strptime(str(fechaHorario) + ' ' + str(record[15+horarioSet] + ':00'),"%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                            diaHorario.horaFinalSabado = datetime.strptime(str(fechaHorario) + ' ' + str(record[16+horarioSet]  + ':00'),"%Y-%m-%d %H:%M:%S") + timedelta(hours=6)

                    elif fechaHorario.weekday() == 6:
                        if record[17 + horarioSet] != None:
                            diaHorario.horaInicioDomingo = datetime.strptime(str(fechaHorario) + ' ' + str(record[17+horarioSet] + ':00'),"%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                            diaHorario.horaFinalDomingo = datetime.strptime(str(fechaHorario) + ' ' + str(record[18+horarioSet]  + ':00'),"%Y-%m-%d %H:%M:%S") + timedelta(hours=6)

                    fechaHorario += step

                semanasWhile += 1
                fechaHorarioT += timedelta(weeks=1)

        if len(empleadosNoEncontradosList) > 0:
            template_id = self.env.ref('horario.email_empleado_no_encontrado').id
            template = self.env['mail.template'].browse(template_id)
            email_values = {'email_to': self.env.user.email,
                            'email_from': ICPSudo.get_param('nomina.correoEnvio'),
                            'subject': 'Horarios no Cargados'
                            }
            template.with_context(datosCorreo=empleadosNoEncontradosList).send_mail(self.id, email_values=email_values, force_send=True)

    def verificar_horario_empleado(self):
        wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.documentoHorario)),read_only=True)

        ws = wb.active
        verificar = False

        empleadosNoEncontradosList = []
        for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=None, values_only=True):

            empleados = self.env['hr.employee'].search(['&',('active','=',True),('department_id.name', '!=','Docentes'),('department_id','!=',"Inactivos")])

            empleado = list(filter(lambda x: (str(x.identification_id).replace("-", "") == str(record[3]).replace("-", "") ),empleados))

            if len(empleado) <= 0:
                raise ValidationError('La identificacion ' + str(record[3]) + '\n No se pudo asociar a ningun empleado')
                continue

            contrato = self.env['contrato.empleado'].search([('empleado_id', '=', empleado[0].id)])

            horas_semanales_horario_1 = timedelta(minutes=0)
            horas_semanales_horario_2 = timedelta(minutes=0)
            horarioSet = 0
            horasAlmuerzo = 0
            fechaHorario = record[4].date()

            if record[5 + horarioSet] != None:
                horasAlmuerzo += int(contrato.tiempoAlmuerzo)
                horaInicioLunes = datetime.strptime(str(fechaHorario) + ' ' + str(record[5 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                horaFinalLunes = datetime.strptime(str(fechaHorario) + ' ' + str(record[6 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                horas_semanales_horario_1 = (horaFinalLunes - horaInicioLunes)

            if record[7 + horarioSet] != None:
                horasAlmuerzo += int(contrato.tiempoAlmuerzo)
                horaInicioMartes = datetime.strptime(str(fechaHorario) + ' ' + str(record[7 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                horaFinalMartes = datetime.strptime(str(fechaHorario) + ' ' + str(record[8 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                horas_semanales_horario_1 += (horaFinalMartes - horaInicioMartes)

            if record[9 + horarioSet] != None:
                horasAlmuerzo += int(contrato.tiempoAlmuerzo)
                horaInicioMiercoles = datetime.strptime(str(fechaHorario) + ' ' + str(record[9  + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                horaFinalMiercoles =  datetime.strptime(str(fechaHorario) + ' ' + str(record[10 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                horas_semanales_horario_1 += (horaFinalMiercoles - horaInicioMiercoles)

            if record[11 + horarioSet] != None:
                horasAlmuerzo += int(contrato.tiempoAlmuerzo)
                horaInicioJueves = datetime.strptime(str(fechaHorario) + ' ' + str(record[11 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                horaFinalJueves = datetime.strptime(str(fechaHorario) + ' ' + str(record[12 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                horas_semanales_horario_1 += (horaFinalJueves - horaInicioJueves)

            if record[13 + horarioSet] != None:
                horasAlmuerzo += int(contrato.tiempoAlmuerzo)
                horaInicioViernes = datetime.strptime(str(fechaHorario) + ' ' + str(record[13 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                horaFinalViernes = datetime.strptime(str(fechaHorario) + ' ' + str(record[14 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                horas_semanales_horario_1 += (horaFinalViernes - horaInicioViernes)

            if record[15 + horarioSet] != None:
                horaInicioSabado = datetime.strptime(str(fechaHorario) + ' ' + str(record[15 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                horaFinalSabado = datetime.strptime(str(fechaHorario) + ' ' + str(record[16 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                horas_semanales_horario_1 += (horaFinalSabado - horaInicioSabado)

            if record[17 + horarioSet] != None:
                horaInicioDomingo = datetime.strptime(str(fechaHorario) + ' ' + str(record[17 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                horaFinalDomingo = datetime.strptime(str(fechaHorario) + ' ' + str(record[18 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                horas_semanales_horario_1 += (horaFinalDomingo - horaInicioDomingo)

            horas_semanales_horario_1 = (((horas_semanales_horario_1.total_seconds() / 60) - horasAlmuerzo) / 60 )
            if horas_semanales_horario_1 >= contrato.jornadaLaboral:
                verificar = True
            else:
                verificar = False
                raise ValidationError('El horario del empleado ' + empleado[0].name+' con la identificacion '+str(record[3])+' cuenta con '+ str(horas_semanales_horario_1)+'h las cuales no cunple con las horas asignadas en el contrato las cuales son ' + str(contrato.jornadaLaboral))

            if record[2] == 'SI' :
                horarioSet = 14
                horasAlmuerzo = 0
                if record[5 + horarioSet] != None:
                    horasAlmuerzo += int(contrato.tiempoAlmuerzo)
                    horaInicioLunes = datetime.strptime(str(fechaHorario) + ' ' + str(record[5 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                    horaFinalLunes = datetime.strptime(str(fechaHorario) + ' ' + str(record[6 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                    horas_semanales_horario_2 = (horaFinalLunes - horaInicioLunes)

                if record[7 + horarioSet] != None:
                    horasAlmuerzo += int(contrato.tiempoAlmuerzo)
                    horaInicioMartes = datetime.strptime(str(fechaHorario) + ' ' + str(record[7 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                    horaFinalMartes = datetime.strptime(str(fechaHorario) + ' ' + str(record[8 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                    horas_semanales_horario_2 += (horaFinalMartes - horaInicioMartes)

                if record[9 + horarioSet] != None:
                    horasAlmuerzo += int(contrato.tiempoAlmuerzo)
                    horaInicioMiercoles = datetime.strptime(str(fechaHorario) + ' ' + str(record[9 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                    horaFinalMiercoles = datetime.strptime(str(fechaHorario) + ' ' + str(record[10 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                    horas_semanales_horario_2 += (horaFinalMiercoles - horaInicioMiercoles)

                if record[11 + horarioSet] != None:
                    horasAlmuerzo += int(contrato.tiempoAlmuerzo)
                    horaInicioJueves = datetime.strptime(str(fechaHorario) + ' ' + str(record[11 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                    horaFinalJueves = datetime.strptime(str(fechaHorario) + ' ' + str(record[12 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                    horas_semanales_horario_2 += (horaFinalJueves - horaInicioJueves)

                if record[13 + horarioSet] != None:
                    horasAlmuerzo += int(contrato.tiempoAlmuerzo)
                    horaInicioViernes = datetime.strptime(str(fechaHorario) + ' ' + str(record[13 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                    horaFinalViernes = datetime.strptime(str(fechaHorario) + ' ' + str(record[14 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                    horas_semanales_horario_2 += (horaFinalViernes - horaInicioViernes)

                if record[15 + horarioSet] != None:
                    horaInicioSabado = datetime.strptime(str(fechaHorario) + ' ' + str(record[15 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                    horaFinalSabado = datetime.strptime(str(fechaHorario) + ' ' + str(record[16 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                    horas_semanales_horario_2 += (horaFinalSabado - horaInicioSabado)

                if record[17 + horarioSet] != None:
                    horaInicioDomingo = datetime.strptime(str(fechaHorario) + ' ' + str(record[17 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                    horaFinalDomingo = datetime.strptime(str(fechaHorario) + ' ' + str(record[18 + horarioSet] + ':00'), "%Y-%m-%d %H:%M:%S") + timedelta(hours=6)
                    horas_semanales_horario_2 += (horaFinalDomingo - horaInicioDomingo)

                horas_semanales_horario_2 = (((horas_semanales_horario_2.total_seconds() / 60) - horasAlmuerzo) / 60)
                if horas_semanales_horario_2 >= contrato.jornadaLaboral:
                    verificar = True
                else:
                    verificar = False
                    raise ValidationError('El horario 2 del empleado ' + empleado[0].name + ' con la identificacion ' + str(record[3]) + ' cuenta con ' + str(horas_semanales_horario_2) + 'h las cuales no cunple con las horas asignadas en el contrato las cuales son ' + str(contrato.jornadaLaboral))

        if verificar:
            self.cargar_horario_empleado()

